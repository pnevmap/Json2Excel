import json
from openpyxl import Workbook
menuCourses = json.loads(open("menu.json").read())

wb = Workbook()
ws = wb.create_sheet("categories")
for menuCourse in menuCourses:
    menuCourseName = menuCourse['menuName']
    row = [menuCourseName]
    sheet = wb.get_sheet_by_name('categories')
    sheet.append(row)
    ws = wb.create_sheet(title=menuCourseName)
    for menuItem in menuCourse['menuItems']:
        row = [menuItem['menuItemName'], menuItem['menuItemDescription']]
        ws.append(row)
wb.save('menu.xlsx')